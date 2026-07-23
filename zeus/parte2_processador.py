import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------------------------------------------------------
# Layout de saída: 14 colunas (5 Livro | sep | 7 Book | Diferença)
# ---------------------------------------------------------------------------
NOMES_COLUNAS_SAIDA = [
    'Data Emissão', 'Nota', 'Fornecedor', 'CNPJ/CPF/CEI/CAEPF', 'Valor Contábil',
    ' ',
    '#Boleta', 'Data Fim', 'Grupo', 'Razão Social', 'CNPJ', 'Valor Total', 'NFE',
    'Diferença do Bloco'
]

NOMES_COLUNAS_SAIDA_SAIDA = [
    'Data Emissão', 'Nota', 'Cliente', 'CNPJ/CPF/CEI/CAEPF', 'Valor Contábil',
    ' ',
    '#Boleta', 'Data Fim', 'Grupo', 'Razão Social', 'CNPJ', 'Valor Total', 'NFE',
    'Diferença do Bloco'
]

COLUNAS_LIVRO = {
    "nome":        "Fornecedor",
    "cnpj":        "CNPJ/CPF/CEI/CAEPF",
    "valor":       "Valor Contábil",
    "nota":        "Nota",
    "data_emissao": "Data Emissão",
}

COLUNAS_LIVRO_SAIDA = {
    "nome":        "Cliente",
    "cnpj":        "CNPJ/CPF/CEI/CAEPF",
    "valor":       "Valor Contábil",
    "nota":        "Nota",
    "data_emissao": "Data Emissão",
}

COLUNAS_BOOK = {
    "cv":        "C/V",           # usado para split Compra/Venda
    "nome":      "Razão Social",
    "cnpj":      "CNPJ",
    "valor":     "Valor Total",
    "data_fim":  "Data Fim",
    "boleta":    "#Boleta",
    "grupo":     "Grupo",
    "nfe":       "NFE",
}

# Colunas de valor (1-based): Valor Contábil=5, Valor Total=12, Diferença=14
_VALUE_COLS = [5, 12, 14]
# Colunas de data (1-based): Data Emissão=1, Data Fim=8
_DATE_COLS = [1, 8]

_N_COLS = 14  # total de colunas de saída


# ---------------------------------------------------------------------------
# Leitura do Book Zeus (aba única)
# ---------------------------------------------------------------------------

def consolidar_book(arquivo_bytes):
    """Lê a aba do Book Zeus e separa em Compra e Venda pela coluna C/V."""
    xls = pd.ExcelFile(arquivo_bytes)

    # Abas de book: todas a partir da 3ª (índice >= 2), independente do nome.
    # As duas primeiras abas são sempre Entrada e Saída, nessa ordem.
    abas_book = xls.sheet_names[2:]
    if not abas_book:
        print("[Zeus] Nenhuma aba de book encontrada.")
        return pd.DataFrame(), pd.DataFrame()

    frames = []
    cv_col = COLUNAS_BOOK['cv']

    for aba in abas_book:
        try:
            df_raw = pd.read_excel(xls, sheet_name=aba, header=None, nrows=3)
            header_row = 0
            for i in range(min(3, len(df_raw))):
                row_str = ' '.join(str(v) for v in df_raw.iloc[i].tolist())
                if 'C/V' in row_str or 'Razão Social' in row_str:
                    header_row = i
                    break
            df = pd.read_excel(xls, sheet_name=aba, skiprows=header_row, header=0)
            frames.append(df)
            if cv_col in df.columns:
                tipos = df[cv_col].astype(str).str.strip()
                n_compra = (tipos == 'Compra').sum()
                n_venda  = (tipos == 'Venda').sum()
                print(f"[Zeus] '{aba}': {n_compra} Compras, {n_venda} Vendas ({len(df)} linhas total)")
            else:
                print(f"[Zeus] '{aba}': {len(df)} linhas (coluna C/V não encontrada)")
                print(f"[Zeus DEBUG] colunas encontradas: {list(df.columns[:10])}")
        except Exception as e:
            print(f"[Zeus] Aviso: não foi possível ler a aba '{aba}': {e}")

    if not frames:
        return pd.DataFrame(), pd.DataFrame()

    df_all = pd.concat(frames, ignore_index=True)

    if cv_col not in df_all.columns:
        print(f"[Zeus] Coluna '{cv_col}' não encontrada. Colunas: {list(df_all.columns)}")
        return pd.DataFrame(), pd.DataFrame()

    df_all[cv_col] = df_all[cv_col].astype(str).str.strip()
    df_compra = df_all[df_all[cv_col] == 'Compra'].copy().reset_index(drop=True)
    df_venda  = df_all[df_all[cv_col] == 'Venda'].copy().reset_index(drop=True)

    print(f"[Zeus] TOTAL consolidado: {len(df_compra)} Compras, {len(df_venda)} Vendas.")
    return df_compra, df_venda


# ---------------------------------------------------------------------------
# Funções de processamento (idênticas ao Thunders, adaptadas para Zeus)
# ---------------------------------------------------------------------------

def _preparar_dataframe(df_raw, col_config, extra_cols_keys):
    """Padroniza um DataFrame para conciliação."""
    colunas_essenciais = [col_config['cnpj'], col_config['nome'], col_config['valor']]
    colunas_extras = [col for col in [col_config.get(k) for k in extra_cols_keys] if col]
    colunas_para_extrair = list(dict.fromkeys(colunas_essenciais + colunas_extras))
    colunas_existentes = [col for col in colunas_para_extrair if col in df_raw.columns]

    if not all(c in df_raw.columns for c in colunas_essenciais):
        faltando = [c for c in colunas_essenciais if c not in df_raw.columns]
        print(f"[Zeus] Colunas essenciais não encontradas: {faltando}")
        return pd.DataFrame()

    df = df_raw[colunas_existentes].copy()
    df.dropna(subset=[col_config['cnpj'], col_config['valor']], inplace=True)

    df.rename(columns={
        col_config['cnpj']:  'CNPJ_PADRAO',
        col_config['nome']:  'NOME_PADRAO',
        col_config['valor']: 'VALOR_PADRAO',
    }, inplace=True)

    df['CNPJ_PADRAO'] = (
        df['CNPJ_PADRAO'].astype(str).str.strip()
        .str.replace(r'\.0$', '', regex=True)
        .str.replace(r'[.\-/]', '', regex=True)
        .str.zfill(14)
    )
    df['valor_arredondado'] = pd.to_numeric(df['VALOR_PADRAO'], errors='coerce').round(2)
    df.dropna(subset=['valor_arredondado'], inplace=True)
    df['indice_original'] = df.index

    return df


def _criar_resultado_final(merged_data, config):
    """Monta o DataFrame de saída com layout Zeus (14 colunas)."""
    nomes_saida   = config.get('nomes_colunas_saida', NOMES_COLUNAS_SAIDA)
    colunas_livro = config.get('colunas_livro', COLUNAS_LIVRO)

    if merged_data.empty:
        return pd.DataFrame(columns=nomes_saida)

    merged_data = merged_data.copy()
    merged_data['Nome_Ordenacao'] = merged_data['NOME_PADRAO_esq'].fillna(merged_data['NOME_PADRAO_dir'])
    merged_data = merged_data.sort_values(
        by=['Nome_Ordenacao', 'chave_agrupamento', 'chave_emparelhamento']
    ).reset_index(drop=True)
    merged_data['diferenca_bloco'] = np.nan

    def safe_get(col):
        return merged_data[col] if col and col in merged_data.columns else pd.Series(index=merged_data.index)

    col_data_esq   = f"{colunas_livro.get('data_emissao')}_esq"
    col_nota_esq   = f"{colunas_livro.get('nota')}_esq"
    col_boleta_dir = f"{COLUNAS_BOOK.get('boleta')}_dir"
    col_data_fim   = f"{COLUNAS_BOOK.get('data_fim')}_dir"
    col_grupo_dir  = f"{COLUNAS_BOOK.get('grupo')}_dir"
    col_nfe_dir    = f"{COLUNAS_BOOK.get('nfe')}_dir"

    final_data = {
        nomes_saida[0]:  safe_get(col_data_esq),          # Data Emissão
        nomes_saida[1]:  safe_get(col_nota_esq),           # Nota
        nomes_saida[2]:  safe_get('NOME_PADRAO_esq'),      # Fornecedor / Cliente
        nomes_saida[3]:  safe_get('CNPJ_PADRAO_esq'),      # CNPJ Livro
        nomes_saida[4]:  safe_get('VALOR_PADRAO_esq'),     # Valor Contábil
        nomes_saida[5]:  '',                                # Separador
        nomes_saida[6]:  safe_get(col_boleta_dir),         # #Boleta
        nomes_saida[7]:  safe_get(col_data_fim),           # Data Fim
        nomes_saida[8]:  safe_get(col_grupo_dir),          # Grupo
        nomes_saida[9]:  safe_get('NOME_PADRAO_dir'),      # Razão Social
        nomes_saida[10]: safe_get('CNPJ_PADRAO_dir'),      # CNPJ Book
        nomes_saida[11]: safe_get('VALOR_PADRAO_dir'),     # Valor Total
        nomes_saida[12]: safe_get(col_nfe_dir),            # NFE
        nomes_saida[13]: safe_get('diferenca_bloco'),      # Diferença do Bloco
    }

    return pd.DataFrame(final_data)


def _encontrar_melhores_matches(df1, df2, grouping_key_length):
    """Encontra os melhores pares por menor diferença de valor dentro de cada grupo de CNPJ."""
    df1 = df1.copy()
    df2 = df2.copy()
    df1['chave_agrupamento'] = df1['CNPJ_PADRAO'].str[:grouping_key_length]
    df2['chave_agrupamento'] = df2['CNPJ_PADRAO'].str[:grouping_key_length]

    todas_chaves = set(df1['chave_agrupamento'].dropna()).union(set(df2['chave_agrupamento'].dropna()))
    matched_indices_1 = set()
    matched_indices_2 = set()

    for chave in todas_chaves:
        grupo_esq = df1[df1['chave_agrupamento'] == chave]
        grupo_dir = df2[df2['chave_agrupamento'] == chave]

        for _, row_e in grupo_esq.iterrows():
            if row_e['indice_original'] in matched_indices_1:
                continue
            melhor_diferenca = np.inf
            melhor_par_dir_idx = None

            for _, row_d in grupo_dir.iterrows():
                if row_d['indice_original'] in matched_indices_2:
                    continue
                diferenca = abs(row_e['valor_arredondado'] - row_d['valor_arredondado'])
                if diferenca < melhor_diferenca:
                    melhor_diferenca = diferenca
                    melhor_par_dir_idx = row_d['indice_original']

            if melhor_diferenca <= 0.01 and melhor_par_dir_idx is not None:
                matched_indices_1.add(row_e['indice_original'])
                matched_indices_2.add(melhor_par_dir_idx)

    return matched_indices_1, matched_indices_2


def _capturar_transacoes_combinadas(df1, df2, indices_combinados_1, indices_combinados_2, config):
    nomes_saida = config.get('nomes_colunas_saida', NOMES_COLUNAS_SAIDA)
    tc1 = df1[df1['indice_original'].isin(indices_combinados_1)].copy()
    tc2 = df2[df2['indice_original'].isin(indices_combinados_2)].copy()

    if tc1.empty and tc2.empty:
        return pd.DataFrame(columns=nomes_saida)

    grouping_key_length = 8
    tc1['chave_agrupamento'] = tc1['CNPJ_PADRAO'].str[:grouping_key_length]
    tc1['chave_emparelhamento'] = tc1.groupby('chave_agrupamento').cumcount()
    tc2['chave_agrupamento'] = tc2['CNPJ_PADRAO'].str[:grouping_key_length]
    tc2['chave_emparelhamento'] = tc2.groupby('chave_agrupamento').cumcount()

    cols1 = {c: f"{c}_esq" for c in tc1.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    tc1.rename(columns=cols1, inplace=True)
    cols2 = {c: f"{c}_dir" for c in tc2.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    tc2.rename(columns=cols2, inplace=True)

    merged = pd.merge(tc1, tc2, on=['chave_agrupamento', 'chave_emparelhamento'], how='outer')
    if merged.empty:
        return pd.DataFrame(columns=nomes_saida)

    return _criar_resultado_final(merged, config)


def _processar_comparacao(df1, df2, config, indices_combinados_1, indices_combinados_2):
    """Separa sobras, aplica filtro de bloco ≤ 0.01 e retorna (divergentes, iguais)."""
    nomes_saida = config.get('nomes_colunas_saida', NOMES_COLUNAS_SAIDA)
    sobras_1 = df1[~df1['indice_original'].isin(indices_combinados_1)].copy()
    sobras_2 = df2[~df2['indice_original'].isin(indices_combinados_2)].copy()

    transacoes_excluidas = _capturar_transacoes_combinadas(
        df1, df2, indices_combinados_1, indices_combinados_2, config
    )

    if sobras_1.empty and sobras_2.empty:
        return pd.DataFrame(columns=nomes_saida), transacoes_excluidas

    grouping_key_length = config.get('chave_agrupamento_final', 8)

    sobras_1['chave_agrupamento'] = sobras_1['CNPJ_PADRAO'].str[:grouping_key_length]
    sobras_1['chave_emparelhamento'] = sobras_1.groupby('chave_agrupamento').cumcount()
    sobras_2['chave_agrupamento'] = sobras_2['CNPJ_PADRAO'].str[:grouping_key_length]
    sobras_2['chave_emparelhamento'] = sobras_2.groupby('chave_agrupamento').cumcount()

    cols1 = {c: f"{c}_esq" for c in sobras_1.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    sobras_1.rename(columns=cols1, inplace=True)
    cols2 = {c: f"{c}_dir" for c in sobras_2.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    sobras_2.rename(columns=cols2, inplace=True)

    merged_sobras = pd.merge(sobras_1, sobras_2, on=['chave_agrupamento', 'chave_emparelhamento'], how='outer')
    blocos_excluidos = pd.DataFrame()

    if not merged_sobras.empty:
        val_esq = pd.to_numeric(merged_sobras.get('VALOR_PADRAO_esq'), errors='coerce').fillna(0)
        val_dir = pd.to_numeric(merged_sobras.get('VALOR_PADRAO_dir'), errors='coerce').fillna(0)
        merged_sobras = merged_sobras.copy()
        merged_sobras['_val_esq'] = val_esq
        merged_sobras['_val_dir'] = val_dir

        group_sums  = merged_sobras.groupby('chave_agrupamento')[['_val_esq', '_val_dir']].sum()
        group_sums['diferenca'] = group_sums['_val_esq'] - group_sums['_val_dir']

        chaves_excluidas = group_sums[abs(group_sums['diferenca']) <= 0.01].index
        blocos_excluidos = merged_sobras[merged_sobras['chave_agrupamento'].isin(chaves_excluidas)].copy()

        chaves_manter = group_sums[abs(group_sums['diferenca']) > 0.01].index
        merged_sobras = merged_sobras[merged_sobras['chave_agrupamento'].isin(chaves_manter)].copy()

        if merged_sobras.empty:
            merged_sobras = pd.DataFrame(columns=nomes_saida)

    resultado_principal = _criar_resultado_final(merged_sobras, config)
    resultado_blocos_excluidos = (
        _criar_resultado_final(blocos_excluidos, config)
        if not blocos_excluidos.empty
        else pd.DataFrame(columns=nomes_saida)
    )

    todos_excluidos = pd.concat([transacoes_excluidas, resultado_blocos_excluidos], ignore_index=True)
    return resultado_principal, todos_excluidos


def _formatar_aba_final(workbook, config, resultado_final):
    """Cria e formata a aba de resultado com layout Zeus (14 colunas)."""
    nomes_saida   = config.get('nomes_colunas_saida', NOMES_COLUNAS_SAIDA)
    nova_aba_nome = config['nome_aba_saida']
    if nova_aba_nome in workbook.sheetnames:
        del workbook[nova_aba_nome]
    ws = workbook.create_sheet(title=nova_aba_nome)

    cabecalho_display = nomes_saida.copy()
    cabecalho_display[3]  = 'CNPJ'   # CNPJ/CPF Livro
    cabecalho_display[10] = 'CNPJ'   # CNPJ Book (já é 'CNPJ')
    ws.append(cabecalho_display)

    orange_fill  = PatternFill(start_color="FFE46C0A", end_color="FFE46C0A", fill_type="solid")
    blue_fill    = PatternFill(start_color="FF002060", end_color="FF002060", fill_type="solid")
    gray_fill    = PatternFill(start_color="808080",   end_color="808080",   fill_type="solid")
    dk_gray_fill = PatternFill(start_color="D9D9D9",   end_color="D9D9D9",   fill_type="solid")
    total_fill   = PatternFill(start_color="BFBFBF",   end_color="BFBFBF",   fill_type="solid")

    header_font   = Font(name='Calibri', bold=True, italic=True, color="FFFFFF", size=11)
    data_font     = Font(name='Calibri', size=11)
    red_data_font = Font(name='Calibri', size=11, color="FF0000")
    summary_font  = Font(name='Calibri', bold=True, size=11)
    red_font      = Font(name='Calibri', bold=True, size=11, color="FF0000")
    total_font    = Font(name='Calibri', bold=True, size=11)
    red_total     = Font(name='Calibri', bold=True, size=11, color="FF0000")
    center_align  = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 25
    for i, cell in enumerate(ws[1]):
        cell.font      = header_font
        cell.alignment = center_align
        col = i + 1
        if 1 <= col <= 5:
            cell.fill = orange_fill   # Livro
        elif 7 <= col <= 13:
            cell.fill = blue_fill     # Book (cols 7-13)
        elif col == 14:
            cell.fill = gray_fill     # Diferença

    if not resultado_final.empty:
        resultado_final = resultado_final.copy()
        cnpjs = resultado_final[nomes_saida[3]].fillna(resultado_final[nomes_saida[10]])
        grouping_key_length = config.get('chave_agrupamento_final', 8)
        resultado_final['_group_key'] = cnpjs.astype(str).str[:grouping_key_length]

        current_row = 2
        is_gray = False

        for _, group_df in resultado_final.groupby('_group_key', sort=False):
            is_gray = not is_gray
            group_df = group_df.drop(columns=['_group_key'])

            for _, data_row in group_df.iterrows():
                ws.append(data_row.tolist())
                if is_gray:
                    for cell in ws[current_row]:
                        cell.fill = dk_gray_fill
                for col_idx in _VALUE_COLS:
                    cell = ws.cell(row=current_row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        if cell.value < 0:
                            cell.font = red_data_font
                for col_idx in _DATE_COLS:
                    cell = ws.cell(row=current_row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = 'DD/MM/YYYY'
                current_row += 1

            soma_esq  = pd.to_numeric(group_df[nomes_saida[4]],  errors='coerce').sum()
            soma_dir  = pd.to_numeric(group_df[nomes_saida[11]], errors='coerce').sum()
            diferenca = soma_esq - soma_dir

            summary_row = [''] * _N_COLS
            summary_row[4]  = soma_esq  if soma_esq  != 0 else None
            summary_row[11] = soma_dir  if soma_dir  != 0 else None
            summary_row[13] = diferenca
            ws.append(summary_row)

            for cell in ws[current_row]:
                cell.font = summary_font
                if is_gray:
                    cell.fill = dk_gray_fill
            for col_idx in _VALUE_COLS:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.number_format = '#,##0.00'
                if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = red_font
            current_row += 1

    if not resultado_final.empty:
        total_esq  = pd.to_numeric(resultado_final.iloc[:, 4],  errors='coerce').sum()
        total_dir  = pd.to_numeric(resultado_final.iloc[:, 11], errors='coerce').sum()
        total_diff = total_esq - total_dir

        ws.append([])
        total_row = ws.max_row + 1

        ws.cell(row=total_row, column=1, value="TOTAL GERAL").font = total_font

        c1 = ws.cell(row=total_row, column=5,  value=total_esq)
        c1.font = red_total if total_esq < 0 else total_font
        c1.number_format = '#,##0.00'

        c2 = ws.cell(row=total_row, column=12, value=total_dir)
        c2.font = red_total if total_dir < 0 else total_font
        c2.number_format = '#,##0.00'

        c3 = ws.cell(row=total_row, column=14, value=total_diff)
        c3.font = red_total if total_diff < 0 else total_font
        c3.number_format = '#,##0.00'

        for cell in ws[total_row]:
            cell.fill = total_fill

    for i in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if i in _VALUE_COLS and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    if cell.value < 0 and cell.row > 1:
                        cell.font = red_data_font
                if i in _DATE_COLS and cell.value is not None and cell.row > 1:
                    cell.number_format = 'DD/MM/YYYY'
                char_count = len(str(cell.value or ""))
                if char_count > max_length:
                    max_length = char_count
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3


def _executar_base(workbook, config, df_livro_raw, df_book_raw, usar_exclusao_parcial):
    """Lógica comum de execução para os dois modos de comparação."""
    if workbook is None:
        return workbook
    print(f"\n--- INICIANDO {config['nome_processo']} ---")
    try:
        df_livro    = df_livro_raw.copy()
        colunas_livro = config.get('colunas_livro', COLUNAS_LIVRO)
        cutoff_date_str = config.get('data_corte')
        date_col = colunas_livro.get('data_emissao')

        if cutoff_date_str and date_col and date_col in df_livro.columns:
            try:
                cutoff = pd.to_datetime(cutoff_date_str)
                df_livro[date_col] = pd.to_datetime(df_livro[date_col], errors='coerce')
                antes = len(df_livro)
                df_livro = df_livro[df_livro[date_col] < cutoff].copy()
                print(f"[{config['nome_processo']}] Filtro de data: {antes - len(df_livro)} linhas removidas.")
            except Exception as e:
                print(f"Aviso: filtro de data falhou: {e}")

        df1 = _preparar_dataframe(df_livro,    colunas_livro, ['nota', 'data_emissao'])
        df2 = _preparar_dataframe(df_book_raw, COLUNAS_BOOK,  ['boleta', 'data_fim', 'grupo', 'nfe'])

        if df1.empty and df2.empty:
            _formatar_aba_final(workbook, config, pd.DataFrame(columns=NOMES_COLUNAS_SAIDA))
            return workbook

        if usar_exclusao_parcial:
            indices_1, indices_2 = _encontrar_melhores_matches(
                df1, df2, config.get('chave_agrupamento_final', 8)
            )
        else:
            merged = pd.merge(df1, df2, on='CNPJ_PADRAO', how='outer', suffixes=('_1', '_2'))
            condicao = abs(merged['valor_arredondado_1'] - merged['valor_arredondado_2']) <= 0.01
            indices_1 = set(merged.loc[condicao, 'indice_original_1'].dropna())
            indices_2 = set(merged.loc[condicao, 'indice_original_2'].dropna())

        resultado_final, resultado_excluidos = _processar_comparacao(df1, df2, config, indices_1, indices_2)
        _formatar_aba_final(workbook, config, resultado_final)

        if usar_exclusao_parcial:
            config_excl = config.copy()
            config_excl['nome_aba_saida'] = config['nome_aba_saida'] + ' - ='
            _formatar_aba_final(workbook, config_excl, resultado_excluidos)
            print(f"[{config['nome_processo']}] Aba de iguais criada: {config_excl['nome_aba_saida']}")

        print(f"[{config['nome_processo']}] Concluído.")
        return workbook
    except Exception as e:
        import traceback
        print(f"Erro em {config['nome_processo']}: {e}")
        traceback.print_exc()
        return workbook


def executar_comparacao_zeus(workbook, config, df_livro_raw, df_book_raw):
    """Comparação padrão lado a lado (agrupamento por 12 dígitos do CNPJ)."""
    return _executar_base(workbook, config, df_livro_raw, df_book_raw, usar_exclusao_parcial=False)


def executar_exclusao_parcial_zeus(workbook, config, df_livro_raw, df_book_raw):
    """Comparação com exclusão parcial (agrupamento por 8 dígitos do CNPJ)."""
    return _executar_base(workbook, config, df_livro_raw, df_book_raw, usar_exclusao_parcial=True)
