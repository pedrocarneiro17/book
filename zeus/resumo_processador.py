import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


def get_totals_from_sheet(workbook_obj, sheet_name):
    """Lê a linha 'TOTAL GERAL' e extrai os totais do Livro (col 5) e Book (col 12)."""
    try:
        ws = workbook_obj[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if row[0] == "TOTAL GERAL":
                total_livro = pd.to_numeric(row[4],  errors='coerce')   # col 5  → Valor Contábil
                total_book  = pd.to_numeric(row[11], errors='coerce')   # col 12 → Valor Total
                return (
                    total_livro if not pd.isna(total_livro) else 0,
                    total_book  if not pd.isna(total_book)  else 0,
                )
        return 0, 0
    except KeyError:
        print(f"[Zeus Resumo] Aba '{sheet_name}' não encontrada — sem pares iguais nesta direção.")
        return 0, 0


def _apply_styles(ws):
    """Aplica formatação na aba de resumo Zeus."""
    header_font  = Font(name='Calibri', bold=True, color="FFFFFF", size=14)
    section_font = Font(name='Calibri', bold=True, color="000000", size=12)
    label_font   = Font(name='Calibri', bold=True, size=11)
    value_font   = Font(name='Calibri', size=11)
    red_value    = Font(name='Calibri', size=11, color="FF0000")
    diff_label   = Font(name='Calibri', bold=True, italic=True, size=11)
    diff_value   = Font(name='Calibri', bold=True, italic=True, size=11)
    red_diff     = Font(name='Calibri', bold=True, italic=True, size=11, color="FF0000")
    data_font    = Font(name='Calibri', size=11)

    orange_fill  = PatternFill(start_color="FFE46C0A", end_color="FFE46C0A", fill_type="solid")
    dk_gray_fill = PatternFill(start_color="D9D9D9",   end_color="D9D9D9",   fill_type="solid")

    ws['B2'].fill = orange_fill
    ws['B2'].font = header_font
    ws.row_dimensions[2].height = 30

    # Seção ENTRADA (linhas 4-14)
    for cell in ws['B4:D4'][0]:
        cell.fill = dk_gray_fill
    ws['B4'].font = section_font
    ws['D4'].font = label_font

    for row_idx in [5, 6, 9, 10]:
        ws.cell(row=row_idx, column=2).font = value_font
        cell = ws.cell(row=row_idx, column=3)
        cell.font = red_value if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else value_font
        cell.number_format = '#,##0.00'

    for row_idx in [7, 11]:
        ws.cell(row=row_idx, column=2).font = label_font
        cell = ws.cell(row=row_idx, column=3)
        cell.font = red_value if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else label_font
        cell.number_format = '#,##0.00'

    for row_idx in [13, 14]:
        ws.cell(row=row_idx, column=2).font = diff_label
        cell = ws.cell(row=row_idx, column=3)
        cell.font = red_diff if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value
        cell.number_format = '#,##0.00'

    # Seção SAÍDA (linhas 16-26)
    for cell in ws['B16:D16'][0]:
        cell.fill = dk_gray_fill
    ws['B16'].font = section_font
    ws['D16'].font = label_font

    for row_idx in [17, 18, 21, 22]:
        ws.cell(row=row_idx, column=2).font = value_font
        cell = ws.cell(row=row_idx, column=3)
        cell.font = red_value if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else value_font
        cell.number_format = '#,##0.00'

    for row_idx in [19, 23]:
        ws.cell(row=row_idx, column=2).font = label_font
        cell = ws.cell(row=row_idx, column=3)
        cell.font = red_value if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else label_font
        cell.number_format = '#,##0.00'

    for row_idx in [25, 26]:
        ws.cell(row=row_idx, column=2).font = diff_label
        cell = ws.cell(row=row_idx, column=3)
        cell.font = red_diff if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value
        cell.number_format = '#,##0.00'

    for row_idx in [5, 17]:
        cell = ws.cell(row=row_idx, column=4)
        cell.font = data_font
        if cell.value:
            cell.number_format = 'DD/MM/YYYY'

    for col_letter in ['B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 28


def criar_aba_resumo_zeus(workbook, data_compras, data_vendas):
    """Cria a aba de Resumo para o modelo Zeus."""
    print("\n--- INICIANDO CRIAÇÃO DA ABA DE RESUMO ZEUS ---")

    ws = workbook.create_sheet("Resumo")

    livro_ent_div,    book_ent_div    = get_totals_from_sheet(workbook, "Livro x Book Entrada")
    livro_ent_iguais, book_ent_iguais = get_totals_from_sheet(workbook, "Livro x Book Entrada - =")
    total_livro_entrada = livro_ent_div + livro_ent_iguais
    total_book_entrada  = book_ent_div  + book_ent_iguais
    diff_ent_div    = livro_ent_div    - book_ent_div
    diff_ent_iguais = livro_ent_iguais - book_ent_iguais

    livro_sai_div,    book_sai_div    = get_totals_from_sheet(workbook, "Livro x Book Saída")
    livro_sai_iguais, book_sai_iguais = get_totals_from_sheet(workbook, "Livro x Book Saída - =")
    total_livro_saida = livro_sai_div + livro_sai_iguais
    total_book_saida  = book_sai_div  + book_sai_iguais
    diff_sai_div    = livro_sai_div    - book_sai_div
    diff_sai_iguais = livro_sai_iguais - book_sai_iguais

    ws.merge_cells('B2:D2')
    ws['B2'] = "Conciliação Livro Fiscal x Book Zeus"

    ws['B4'] = "ENTRADA"
    ws['D4'] = "Data do filtro"
    ws['B5'] = "Livro Fiscal divergentes"
    ws['C5'] = livro_ent_div
    ws['D5'] = pd.to_datetime(data_compras, errors='coerce') if data_compras else None
    ws['B6'] = "Livro Fiscal iguais"
    ws['C6'] = livro_ent_iguais
    ws['B7'] = "TOTAL DO LIVRO"
    ws['C7'] = total_livro_entrada
    ws['B8'] = ""
    ws['B9']  = "Book divergentes"
    ws['C9']  = book_ent_div
    ws['B10'] = "Book iguais"
    ws['C10'] = book_ent_iguais
    ws['B11'] = "TOTAL DO BOOK"
    ws['C11'] = total_book_entrada
    ws['B12'] = ""
    ws['B13'] = "Divergentes"
    ws['C13'] = diff_ent_div
    ws['B14'] = "Iguais"
    ws['C14'] = diff_ent_iguais

    ws['B16'] = "SAÍDA"
    ws['D16'] = "Data do filtro"
    ws['B17'] = "Livro Fiscal divergentes"
    ws['C17'] = livro_sai_div
    ws['D17'] = pd.to_datetime(data_vendas, errors='coerce') if data_vendas else None
    ws['B18'] = "Livro Fiscal iguais"
    ws['C18'] = livro_sai_iguais
    ws['B19'] = "TOTAL DO LIVRO"
    ws['C19'] = total_livro_saida
    ws['B20'] = ""
    ws['B21'] = "Book divergentes"
    ws['C21'] = book_sai_div
    ws['B22'] = "Book iguais"
    ws['C22'] = book_sai_iguais
    ws['B23'] = "TOTAL DO BOOK"
    ws['C23'] = total_book_saida
    ws['B24'] = ""
    ws['B25'] = "Divergentes"
    ws['C25'] = diff_sai_div
    ws['B26'] = "Iguais"
    ws['C26'] = diff_sai_iguais

    _apply_styles(ws)
    print("[Zeus Resumo] Aba de Resumo criada com sucesso.")
    return workbook
