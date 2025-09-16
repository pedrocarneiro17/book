import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy
import numpy as np

def copy_cell_style(source_cell, target_cell):
    """Copia o estilo completo de uma célula para outra."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def _formatar_aba_divergencia(ws_nova, df_resultado):
    """Aplica o layout profissional na nova aba de divergências."""
    total_font = Font(bold=True, size=12)
    red_total_font = Font(bold=True, size=12, color="FF0000")
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    data_font = Font(name='Calibri', size=11)
    red_data_font = Font(name='Calibri', size=11, color="FF0000")
    
    if not df_resultado.empty:
        total_v1 = pd.to_numeric(df_resultado['Valor Ajustado 1'], errors='coerce').sum()
        total_v2 = pd.to_numeric(df_resultado['Valor Ajustado 2'], errors='coerce').sum()
        
        # --- AJUSTE: INVERTIDA A ORDEM DA SUBTRAÇÃO ---
        total_diff = total_v2 - total_v1

        total_row_index = ws_nova.max_row + 2
        ws_nova.cell(row=total_row_index, column=1, value="TOTAL").font = total_font
        for col_idx, total_val in [(7, total_v1), (8, total_v2), (9, total_diff)]:
            cell = ws_nova.cell(row=total_row_index, column=col_idx, value=total_val)
            cell.font = red_total_font if total_val < 0 else total_font
            cell.number_format = '#,##0.00'
        for cell in ws_nova[total_row_index]:
            cell.fill = total_fill

    total_row_index_check = ws_nova.max_row if df_resultado.empty else total_row_index
    for col in ws_nova.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for i, cell in enumerate(col):
            if i > 0 and cell.row != total_row_index_check:
                cell.font = data_font
                if cell.column in [7, 8, 9] and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    if cell.value < 0:
                        cell.font = red_data_font
            try:
                if isinstance(cell.value, (int, float)) and cell.column in [7, 8, 9]:
                    cell_text = f'{cell.value:,.2f}'  # Remove R$ for width calculation
                else:
                    cell_text = str(cell.value or "")
                char_count = len(cell_text)
                if cell.font and cell.font.size > 11:
                    char_count = char_count * (cell.font.size / 11)
                if char_count > max_length:
                    max_length = char_count
            except: pass
        adjusted_width = (max_length + 3)
        ws_nova.column_dimensions[column_letter].width = adjusted_width

def processa_comparacao_de_abas_p1(workbook, df_principal, df_secundaria, nome_aba_principal, colunas_p1):
    print(f"\n--- [Parte 1] Processando: '{nome_aba_principal}' e sua correspondente ---")
    
    num_contrato_col = colunas_p1['num_contrato']
    valor_col = colunas_p1['valor']
    
    if num_contrato_col not in df_principal.columns or num_contrato_col not in df_secundaria.columns:
        print(f"Aviso: Coluna '{num_contrato_col}' não encontrada. Pulando Parte 1.")
        return
        
    df_principal.dropna(subset=[num_contrato_col, valor_col], inplace=True)
    df_secundaria.dropna(subset=[num_contrato_col, valor_col], inplace=True)
    df_principal[valor_col] = pd.to_numeric(df_principal[valor_col], errors='coerce').round(2)
    df_secundaria[valor_col] = pd.to_numeric(df_secundaria[valor_col], errors='coerce').round(2)
    df_principal.dropna(subset=[valor_col], inplace=True)
    df_secundaria.dropna(subset=[valor_col], inplace=True)

    contratos_principais = set(df_principal[num_contrato_col].unique())
    df_secundaria['MCP'] = np.where(df_secundaria[num_contrato_col].isin(contratos_principais), 'N', 'S')
    sheet_secundaria = workbook[df_secundaria.attrs['nome_aba']]
    df_secundaria_final = df_secundaria.sort_values(by=colunas_p1['nome']).reset_index(drop=True)
    for row in sheet_secundaria.iter_rows(min_row=2):
        for cell in row: cell.value = None
    rows = dataframe_to_rows(df_secundaria_final, index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            sheet_secundaria.cell(row=r_idx, column=c_idx, value=value)
    mcp_col_index = df_secundaria_final.columns.get_loc('MCP') + 1
    header_cell_mcp = sheet_secundaria.cell(row=1, column=mcp_col_index)
    header_cell_mcp.value = 'MCP'
    header_cell_source = sheet_secundaria.cell(row=1, column=mcp_col_index - 1)
    copy_cell_style(header_cell_source, header_cell_mcp)
    for row in sheet_secundaria.iter_rows(min_row=2, max_row=sheet_secundaria.max_row):
        copy_cell_style(row[mcp_col_index - 2], row[mcp_col_index - 1])

    df_p = df_principal.copy()
    df_s = df_secundaria.drop(columns=['MCP']).copy()
    df_p['pairing_key'] = df_p.groupby(num_contrato_col).cumcount()
    df_s['pairing_key'] = df_s.groupby(num_contrato_col).cumcount()
    merged_df = pd.merge(df_p, df_s, on=[num_contrato_col, 'pairing_key'], how='outer', suffixes=('_1', '_2'), indicator=True)
    
    merged_df['valor1_num'] = pd.to_numeric(merged_df[f"{valor_col}_1"], errors='coerce').round(2)
    merged_df['valor2_num'] = pd.to_numeric(merged_df[f"{valor_col}_2"], errors='coerce').round(2)
    condicao_match = (merged_df['valor1_num'] == merged_df['valor2_num'])
    
    divergencias_df = merged_df[~((merged_df['_merge'] == 'both') & condicao_match)].copy()

    if not divergencias_df.empty:
        val1 = pd.to_numeric(divergencias_df[f"{valor_col}_1"], errors='coerce').fillna(0)
        val2 = pd.to_numeric(divergencias_df[f"{valor_col}_2"], errors='coerce').fillna(0)
        
        # --- AJUSTE: INVERTIDA A ORDEM DA SUBTRAÇÃO ---
        divergencias_df['Diferença'] = val2 - val1
        
        final_report_df = pd.DataFrame()
        
        def safe_get(df, col_name, suffix=''):
            full_col_name = f"{col_name}{suffix}"
            return df[full_col_name] if full_col_name in df.columns else pd.Series(index=df.index)

        def coalesce_cols(base_name):
            if not base_name: return pd.Series(index=divergencias_df.index)
            return safe_get(divergencias_df, base_name, '_1').fillna(safe_get(divergencias_df, base_name, '_2'))

        final_report_df['Deal'] = coalesce_cols(colunas_p1.get('deal'))
        final_report_df['Nº Contrato'] = divergencias_df[num_contrato_col]
        final_report_df['Tipo de Operação'] = coalesce_cols(colunas_p1.get('tipo_operacao'))
        final_report_df['Liquidação'] = coalesce_cols(colunas_p1.get('liquidacao'))
        final_report_df['CNPJ'] = coalesce_cols(colunas_p1.get('cnpj'))
        final_report_df['Parte - Contra Banco'] = coalesce_cols(colunas_p1.get('nome'))
        final_report_df['Valor Ajustado 1'] = divergencias_df[f"{valor_col}_1"]
        final_report_df['Valor Ajustado 2'] = divergencias_df[f"{valor_col}_2"]
        final_report_df['Diferença'] = divergencias_df['Diferença']
        
        final_report_df.sort_values(by=['Parte - Contra Banco', 'Nº Contrato'], inplace=True, ignore_index=True)

        if 'Compra' in nome_aba_principal: nova_aba_nome = "Confronto Books Compras"
        else: nova_aba_nome = "Confronto Books Vendas"
        if nova_aba_nome in workbook.sheetnames: del workbook[nova_aba_nome]
        ws_nova = workbook.create_sheet(title=nova_aba_nome)
        for r in dataframe_to_rows(final_report_df, index=False, header=True):
            ws_nova.append(r)
        sheet_principal = workbook[nome_aba_principal]
        if sheet_principal.row_dimensions[1].height:
            ws_nova.row_dimensions[1].height = sheet_principal.row_dimensions[1].height
        for i, col_dim in sheet_principal.column_dimensions.items():
            ws_nova.column_dimensions[i] = copy(col_dim)
        for i in range(1, len(final_report_df.columns) + 1):
            source_cell = sheet_principal.cell(row=1, column=i)
            target_cell = ws_nova.cell(row=1, column=i)
            if source_cell.has_style:
                copy_cell_style(source_cell, target_cell)
        _formatar_aba_divergencia(ws_nova, final_report_df)

    sheet_principal = workbook[nome_aba_principal]
    df_principal_final = df_principal.sort_values(by=colunas_p1['nome']).reset_index(drop=True)
    for row in sheet_principal.iter_rows(min_row=2):
        for cell in row: cell.value = None
    rows_p = dataframe_to_rows(df_principal_final, index=False, header=False)
    for r_idx, row in enumerate(rows_p, 2):
        for c_idx, value in enumerate(row, 1):
            sheet_principal.cell(row=r_idx, column=c_idx, value=value)

def executar_processo_parte1(caminho_entrada, pular_linhas, colunas_p1):
    print("\n=========================================================")
    print("--- INICIANDO EXECUÇÃO DA PARTE 1 ---")
    print("=========================================================")
    try:
        xls = pd.ExcelFile(caminho_entrada)
        nomes_abas = xls.sheet_names
        workbook = openpyxl.load_workbook(caminho_entrada)
        if len(nomes_abas) >= 5:
            nome_aba_3, nome_aba_5 = nomes_abas[2], nomes_abas[4]
            df3 = pd.read_excel(xls, sheet_name=nome_aba_3, skiprows=pular_linhas, header=0)
            df5 = pd.read_excel(xls, sheet_name=nome_aba_5, skiprows=pular_linhas, header=0)
            df5.attrs['nome_aba'] = nome_aba_5
            processa_comparacao_de_abas_p1(workbook, df3, df5, nome_aba_3, colunas_p1)
        else: print("Aviso: Arquivo não possui abas 3 e 5. Pulando esta comparação.")
        if len(nomes_abas) >= 6:
            nome_aba_4, nome_aba_6 = nomes_abas[3], nomes_abas[5]
            df4 = pd.read_excel(xls, sheet_name=nome_aba_4, skiprows=pular_linhas, header=0)
            df6 = pd.read_excel(xls, sheet_name=nome_aba_6, skiprows=pular_linhas, header=0)
            df6.attrs['nome_aba'] = nome_aba_6
            processa_comparacao_de_abas_p1(workbook, df4, df6, nome_aba_4, colunas_p1)
        else: print("\nAviso: Arquivo não possui abas 4 e 6. Pulando esta comparação.")
        print("\n[Parte 1] Processo concluído em memória.")
        return workbook
    except KeyError as e:
        print(f"Erro de Chave (Coluna não encontrada): {e}. Verifique os nomes das colunas no arquivo Excel e na configuração.")
        return None
    except Exception as e: 
        print(f"Ocorreu um erro fatal na Parte 1: {e}")
        return None