import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _apply_summary_styles(ws):
    """Aplica todos os estilos de formatação na aba de resumo."""
    
    # --- Estilos Gerais ---
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=14)
    section_font_compra = Font(name='Calibri', bold=True, color="C00000", size=12)  # Vermelho
    section_font_venda = Font(name='Calibri', bold=True, color="C00000", size=12)  # Vermelho
    section_font_entrada = Font(name='Calibri', bold=True, color="000000", size=12)  # Preto
    section_font_saida = Font(name='Calibri', bold=True, color="000000", size=12)  # Preto
    
    label_font = Font(name='Calibri', bold=True, size=11)
    data_font = Font(name='Calibri', size=11)
    value_font = Font(name='Calibri', size=11)
    red_value_font = Font(name='Calibri', size=11, color="FF0000")
    diff_label_font = Font(name='Calibri', bold=True, italic=True, size=11)
    diff_value_font = Font(name='Calibri', bold=True, italic=True, size=11)
    red_diff_value_font = Font(name='Calibri', bold=True, italic=True, size=11, color="FF0000")

    blue_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE46C0A", end_color="FFE46C0A", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # --- Tabela 1: Conciliação entre Books ---
    ws['B2'].fill = blue_fill
    ws['B2'].font = header_font
    
    for cell in ws['B4:E4'][0]:
        cell.fill = dark_gray_fill
    ws['B4'].font = section_font_compra
    ws['C4'].font = label_font
    ws['D4'].font = label_font
    ws['E4'].font = label_font

    for row_idx in [5, 6, 10, 11]:
        for col_idx in [2, 3]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = value_font
        cell = ws.cell(row=row_idx, column=3)
        cell.number_format = '#,##0.00'
        if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = red_value_font
    
    for row_idx in [7, 12]:
        ws.cell(row=row_idx, column=2).font = diff_label_font
        for col_idx in [3, 4, 5]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = red_diff_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value_font
            cell.number_format = '#,##0.00'
    
    for cell in ws['B9:E9'][0]:
        cell.fill = dark_gray_fill
    ws['B9'].font = section_font_venda
    ws['C9'].font = label_font
    ws['D9'].font = label_font
    ws['E9'].font = label_font

    # --- Tabela 2: Conciliação entre Livro Fiscal x Book ---
    ws['G2'].fill = orange_fill
    ws['G2'].font = header_font

    for cell in ws['G4:I4'][0]:
        cell.fill = dark_gray_fill
    ws['G4'].font = section_font_entrada
    ws['I4'].font = label_font
    
    # Aplicar estilos para as linhas de entrada (linhas 5, 6, 9, 10)
    for row_idx in [5, 6, 9, 10]:
        ws.cell(row=row_idx, column=7).font = value_font
        cell = ws.cell(row=row_idx, column=8)
        cell.font = red_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else value_font
        cell.number_format = '#,##0.00'
    
    # Linhas de TOTAL para entrada (linhas 7 e 11)
    for row_idx in [7, 11]:
        ws.cell(row=row_idx, column=7).font = label_font  # TOTAL em negrito
        cell = ws.cell(row=row_idx, column=8)
        cell.font = red_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else label_font
        cell.number_format = '#,##0.00'
    
    # Linha de diferença da entrada - Divergentes (linha 13)
    ws.cell(row=13, column=7).font = diff_label_font
    cell = ws.cell(row=13, column=8)
    cell.font = red_diff_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value_font
    cell.number_format = '#,##0.00'
    
    # Linha de diferença da entrada - Iguais (linha 14)
    ws.cell(row=14, column=7).font = diff_label_font
    cell = ws.cell(row=14, column=8)
    cell.font = red_diff_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value_font
    cell.number_format = '#,##0.00'
    
    # Cabeçalho da seção SAÍDA (linha 16)
    for cell in ws['G16:I16'][0]:
        cell.fill = dark_gray_fill
    ws['G16'].font = section_font_saida
    ws['I16'].font = label_font
    
    # Aplicar estilos para as linhas de saída (linhas 17, 18, 21, 22)
    for row_idx in [17, 18, 21, 22]:
        ws.cell(row=row_idx, column=7).font = value_font
        cell = ws.cell(row=row_idx, column=8)
        cell.font = red_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else value_font
        cell.number_format = '#,##0.00'
    
    # Linhas de TOTAL para saída (linhas 19 e 23)
    for row_idx in [19, 23]:
        ws.cell(row=row_idx, column=7).font = label_font  # TOTAL em negrito
        cell = ws.cell(row=row_idx, column=8)
        cell.font = red_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else label_font
        cell.number_format = '#,##0.00'
    
    # Linha de diferença da saída - Divergentes (linha 25)
    ws.cell(row=25, column=7).font = diff_label_font
    cell = ws.cell(row=25, column=8)
    cell.font = red_diff_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value_font
    cell.number_format = '#,##0.00'
    
    # Linha de diferença da saída - Iguais (linha 26)
    ws.cell(row=26, column=7).font = diff_label_font
    cell = ws.cell(row=26, column=8)
    cell.font = red_diff_value_font if cell.value and isinstance(cell.value, (int, float)) and cell.value < 0 else diff_value_font
    cell.number_format = '#,##0.00'

    # Formatar as células de data (I5 e I17)
    for row_idx in [5, 17]:
        cell = ws.cell(row=row_idx, column=9)
        cell.font = data_font
        if cell.value and isinstance(cell.value, (str, pd.Timestamp)):
            cell.number_format = 'DD/MM/YYYY'

    # Ajusta largura das colunas
    for col_letter in ['B', 'C', 'D', 'E', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 25

def get_df_from_ws(workbook_obj, sheet_name):
    """Converte uma aba de um workbook openpyxl para um DataFrame pandas."""
    try:
        ws_obj = workbook_obj[sheet_name]
        data = ws_obj.values
        cols = next(data)
        data = list(data)
        return pd.DataFrame(data, columns=cols)
    except KeyError:
        print(f"Aviso: Aba '{sheet_name}' para o resumo não encontrada. Retornando DataFrame vazio.")
        return pd.DataFrame(columns=[f'col_{i}' for i in range(12)])

def get_totals_from_sheet(workbook_obj, sheet_name):
    """Busca a linha de 'TOTAL GERAL' e extrai as somas já calculadas."""
    try:
        ws_obj = workbook_obj[sheet_name]
        total_livro = 0
        total_book = 0
        for row in ws_obj.iter_rows(values_only=True):
            if row[0] == "TOTAL GERAL":
                total_livro = row[4] if row[4] is not None else 0
                total_book = row[10] if row[10] is not None else 0
                
                total_livro_num = pd.to_numeric(total_livro, errors='coerce')
                total_book_num = pd.to_numeric(total_book, errors='coerce')
                
                return total_livro_num if not pd.isna(total_livro_num) else 0, total_book_num if not pd.isna(total_book_num) else 0

        return 0, 0
    except KeyError:
        print(f"Aviso: Aba '{sheet_name}' para o resumo não encontrada.")
        return 0, 0

def criar_aba_resumo(workbook, p1_config, data_compras, data_vendas):
    """Cria e formata a aba de resumo consolidado a partir do workbook em memória."""
    print("\n--- INICIANDO CRIAÇÃO DA ABA DE RESUMO ---")
    
    ws = workbook.create_sheet("Resumo") 
    
    nomes_abas_existentes = workbook.sheetnames
    
    # --- 1. CÁLCULOS PARA A TABELA "CONCILIAÇÃO ENTRE BOOKS" ---
    
    # Compra
    nome_aba_3 = nomes_abas_existentes[2]
    nome_aba_5 = nomes_abas_existentes[4]
    df3 = get_df_from_ws(workbook, nome_aba_3)
    df5 = get_df_from_ws(workbook, nome_aba_5)
    
    soma_v_ajust_3 = pd.to_numeric(df3[p1_config['valor']], errors='coerce').sum()
    soma_v_ajust_5 = pd.to_numeric(df5[p1_config['valor']], errors='coerce').sum()
    contratos_principais_3 = set(df3[p1_config['num_contrato']].unique())
    df5['MCP'] = np.where(df5[p1_config['num_contrato']].isin(contratos_principais_3), 'N', 'S')
    soma_mcp_sim_5 = pd.to_numeric(df5[df5['MCP'] == 'S'][p1_config['valor']], errors='coerce').sum()
    
    diff_books_compra_real = soma_v_ajust_3 - soma_v_ajust_5
    diff_books_compra_display = abs(diff_books_compra_real) 
    diff_final_compra = diff_books_compra_display - soma_mcp_sim_5

    # Venda
    nome_aba_4 = nomes_abas_existentes[3]
    nome_aba_6 = nomes_abas_existentes[5]
    df4 = get_df_from_ws(workbook, nome_aba_4)
    df6 = get_df_from_ws(workbook, nome_aba_6)

    soma_v_ajust_4 = pd.to_numeric(df4[p1_config['valor']], errors='coerce').sum()
    soma_v_ajust_6 = pd.to_numeric(df6[p1_config['valor']], errors='coerce').sum()
    contratos_principais_4 = set(df4[p1_config['num_contrato']].unique())
    df6['MCP'] = np.where(df6[p1_config['num_contrato']].isin(contratos_principais_4), 'N', 'S')
    soma_mcp_sim_6 = pd.to_numeric(df6[df6['MCP'] == 'S'][p1_config['valor']], errors='coerce').sum()
    
    diff_books_venda_real = soma_v_ajust_4 - soma_v_ajust_6
    diff_books_venda_display = abs(diff_books_venda_real)
    diff_final_venda = diff_books_venda_display - soma_mcp_sim_6

    # --- 2. CÁLCULOS PARA "CONCILIAÇÃO ENTRE LIVRO FISCAL X BOOK" ---
    
    # ENTRADA - Dados divergentes (aba principal)
    soma_livro_entrada_div, soma_book_entrada_div = get_totals_from_sheet(workbook, "Livro x Book Entrada")
    
    # ENTRADA - Dados iguais (aba "- =")
    soma_livro_entrada_iguais, soma_book_entrada_iguais = get_totals_from_sheet(workbook, "Livro x Book Entrada - =")
    
    # ENTRADA - Totais
    total_livro_entrada = soma_livro_entrada_div + soma_livro_entrada_iguais
    total_book_entrada = soma_book_entrada_div + soma_book_entrada_iguais
    diff_entrada_divergentes = soma_livro_entrada_div - soma_book_entrada_div
    diff_entrada_iguais = soma_livro_entrada_iguais - soma_book_entrada_iguais
    
    # SAÍDA - Dados divergentes (aba principal)
    soma_livro_saida_div, soma_book_saida_div = get_totals_from_sheet(workbook, "Livro x Book Saída")
    
    # SAÍDA - Dados iguais (aba "- =")
    soma_livro_saida_iguais, soma_book_saida_iguais = get_totals_from_sheet(workbook, "Livro x Book Saída - =")
    
    # SAÍDA - Totais
    total_livro_saida = soma_livro_saida_div + soma_livro_saida_iguais
    total_book_saida = soma_book_saida_div + soma_book_saida_iguais
    diff_saida_divergentes = soma_livro_saida_div - soma_book_saida_div
    diff_saida_iguais = soma_livro_saida_iguais - soma_book_saida_iguais
    
    # --- 3. MONTAGEM DA ABA NO EXCEL ---
    ws.merge_cells('B2:E2')
    ws['B2'] = "Conciliação entre Books"
    ws.row_dimensions[2].height = 30
    
    ws['B4'] = "Compra"
    ws['C4'] = "Valor ajustado"
    ws['D4'] = "MCP (sim)"
    ws['E4'] = "Diferença"
    
    ws['B5'] = nome_aba_3
    ws['C5'] = soma_v_ajust_3
    ws['B6'] = nome_aba_5
    ws['C6'] = soma_v_ajust_5
    
    ws['B7'] = "Diferença"
    ws['C7'] = diff_books_compra_display
    ws['D7'] = soma_mcp_sim_5
    ws['E7'] = diff_final_compra
    
    ws['B9'] = "Venda"
    ws['C9'] = "Valor ajustado"
    ws['D9'] = "MCP (sim)"
    ws['E9'] = "Diferença"
    
    ws['B10'] = nome_aba_4
    ws['C10'] = soma_v_ajust_4
    ws['B11'] = nome_aba_6
    ws['C11'] = soma_v_ajust_6
    
    ws['B12'] = "Diferença"
    ws['C12'] = diff_books_venda_display
    ws['D12'] = soma_mcp_sim_6
    ws['E12'] = diff_final_venda

    ws.merge_cells('G2:I2')
    ws['G2'] = "Conciliação entre Livro Fiscal x Book"
    
    # ENTRADA - Novo padrão igual ao da saída
    ws['G4'] = "ENTRADA"
    ws['I4'] = "Data do filtro"
    
    ws['G5'] = "Livro Fiscal divergentes"
    ws['H5'] = soma_livro_entrada_div
    ws['I5'] = pd.to_datetime(data_compras, errors='coerce') if data_compras else None
    
    ws['G6'] = "Livro Fiscal iguais"
    ws['H6'] = soma_livro_entrada_iguais
    
    ws['G7'] = "TOTAL DO LIVRO"
    ws['H7'] = total_livro_entrada
    
    # Linha em branco
    ws['G8'] = ""
    ws['H8'] = ""
    
    ws['G9'] = "Book divergentes"
    ws['H9'] = soma_book_entrada_div
    
    ws['G10'] = "Book iguais"
    ws['H10'] = soma_book_entrada_iguais
    
    ws['G11'] = "TOTAL DO BOOK"
    ws['H11'] = total_book_entrada
    
    # Linha em branco
    ws['G12'] = ""
    ws['H12'] = ""
    
    ws['G13'] = "Divergentes"
    ws['H13'] = diff_entrada_divergentes
    
    ws['G14'] = "Iguais"
    ws['H14'] = diff_entrada_iguais

    # SAÍDA - Ajustado para nova posição
    ws['G16'] = "SAÍDA"
    ws['I16'] = "Data do filtro"

    ws['G17'] = "Livro Fiscal divergentes"
    ws['H17'] = soma_livro_saida_div
    ws['I17'] = pd.to_datetime(data_vendas, errors='coerce') if data_vendas else None
    
    ws['G18'] = "Livro Fiscal iguais"
    ws['H18'] = soma_livro_saida_iguais
    
    ws['G19'] = "TOTAL DO LIVRO"
    ws['H19'] = total_livro_saida
    
    # Linha em branco
    ws['G20'] = ""
    ws['H20'] = ""
    
    ws['G21'] = "Book divergentes"
    ws['H21'] = soma_book_saida_div
    
    ws['G22'] = "Book iguais"
    ws['H22'] = soma_book_saida_iguais
    
    ws['G23'] = "TOTAL DO BOOK"
    ws['H23'] = total_book_saida
    
    # Linha em branco
    ws['G24'] = ""
    ws['H24'] = ""
    
    ws['G25'] = "Divergentes"
    ws['H25'] = diff_saida_divergentes
    
    ws['G26'] = "Iguais"
    ws['H26'] = diff_saida_iguais

    _apply_summary_styles(ws)
    print("[Resumo] Aba de Resumo criada com sucesso.")
    return workbook