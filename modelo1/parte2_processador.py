import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

def _formatar_aba_final(workbook, config, resultado_final):
    """Aplica o layout profissional na nova aba de resultado, incluindo resumos por bloco."""
    nova_aba_nome = config['nome_aba_saida']
    if nova_aba_nome in workbook.sheetnames:
        del workbook[nova_aba_nome]
    ws_nova = workbook.create_sheet(title=nova_aba_nome)
    
    cabecalho_display = config['nomes_colunas_saida'].copy()
    cabecalho_display[0] = 'CNPJ'
    cabecalho_display[6] = 'CNPJ'
    ws_nova.append(cabecalho_display)
    
    orange_header_fill = PatternFill(start_color="FFE46C0A", end_color="FFE46C0A", fill_type="solid")
    blue_header_fill = PatternFill(start_color="FF002060", end_color="FF002060", fill_type="solid")
    diff_header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_font = Font(name='Calibri', bold=True, italic=True, color="FFFFFF", size=11)
    data_font = Font(name='Calibri', size=11)
    red_data_font = Font(name='Calibri', size=11, color="FF0000")
    summary_font = Font(name='Calibri', bold=True, size=11)
    red_font = Font(name='Calibri', bold=True, size=11, color="FF0000")
    center_alignment = Alignment(horizontal='center', vertical='center')
    dark_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws_nova.row_dimensions[1].height = 25
    header_row = ws_nova[1]
    for i, cell in enumerate(header_row):
        cell.font = header_font
        cell.alignment = center_alignment
        col_index = i + 1
        if col_index <= 5:
            cell.fill = orange_header_fill
        elif col_index >= 7 and col_index <= 11:
            cell.fill = blue_header_fill
        elif col_index == 12:
            cell.fill = diff_header_fill
            
    if not resultado_final.empty:
        # Cria uma chave temporária para agrupar os dados para formatação
        cnpjs = resultado_final[config['nomes_colunas_saida'][0]].fillna(resultado_final[config['nomes_colunas_saida'][6]])
        grouping_key_length = config.get('chave_agrupamento_final', 8)
        resultado_final['temp_group_key'] = cnpjs.str[:grouping_key_length]
        
        current_row_index = 2
        is_gray_group = False

        for group_key, group_df in resultado_final.groupby('temp_group_key', sort=False):
            is_gray_group = not is_gray_group  # Alterna a cor para cada novo grupo

            # Escreve as linhas de dados do grupo, aplicando a cor do bloco
            for _, data_row in group_df.iterrows():
                # Remove a coluna temporária antes de escrever na planilha
                row_data = data_row.drop('temp_group_key').tolist()
                ws_nova.append(row_data)
                if is_gray_group:
                    for cell in ws_nova[current_row_index]:
                        cell.fill = dark_gray_fill
                # Aplica formatação numérica e cor vermelha para valores negativos nos dados
                for col_idx in [5, 11, 12]:  # Colunas de valores (1-based index)
                    cell = ws_nova.cell(row=current_row_index, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        if cell.value < 0:
                            cell.font = red_data_font if cell.font != summary_font else red_font
                current_row_index += 1

            # Calcula os totais e a diferença do bloco
            soma_esq = pd.to_numeric(group_df[config['nomes_colunas_saida'][4]], errors='coerce').sum()
            soma_dir = pd.to_numeric(group_df[config['nomes_colunas_saida'][10]], errors='coerce').sum()
            diferenca_bloco = soma_esq - soma_dir
            
            # Cria a linha de resumo com os totais e a diferença
            summary_row_values = [''] * len(config['nomes_colunas_saida'])
            summary_row_values[4] = soma_esq if soma_esq != 0 else None
            summary_row_values[10] = soma_dir if soma_dir != 0 else None
            summary_row_values[11] = diferenca_bloco  # Coloca a diferença na última coluna
            ws_nova.append(summary_row_values)
            
            # Formata a linha de resumo
            summary_row_obj = ws_nova[current_row_index]
            for cell in summary_row_obj:
                cell.font = summary_font
                if is_gray_group:
                    cell.fill = dark_gray_fill
            
            # Aplica formatação numérica e cor vermelha para valores negativos
            cell_col5 = ws_nova.cell(row=current_row_index, column=5)
            cell_col5.number_format = '#,##0.00'
            if cell_col5.value and cell_col5.value < 0:
                cell.font = red_font
                
            cell_col11 = ws_nova.cell(row=current_row_index, column=11)
            cell_col11.number_format = '#,##0.00'
            if cell_col11.value and cell_col11.value < 0:
                cell.font = red_font
                
            cell_col12 = ws_nova.cell(row=current_row_index, column=12)
            cell_col12.number_format = '#,##0.00'
            if cell_col12.value and cell_col12.value < 0:
                cell.font = red_font
            
            current_row_index += 1

    total_row_index = 0
    if not resultado_final.empty:
        total_col_1 = pd.to_numeric(resultado_final.iloc[:, 4], errors='coerce').sum()
        total_col_2 = pd.to_numeric(resultado_final.iloc[:, 10], errors='coerce').sum()
        total_diff = total_col_1 - total_col_2  # Calcula a diferença total
        total_font_style = Font(name='Calibri', bold=True, size=11)
        total_red_font_style = Font(name='Calibri', bold=True, size=11, color="FF0000")
        total_fill_style = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        ws_nova.append([]) 
        total_row_index = ws_nova.max_row + 1

        total_label_cell = ws_nova.cell(row=total_row_index, column=1, value="TOTAL GERAL")
        total_label_cell.font = total_font_style
        cell_val_1 = ws_nova.cell(row=total_row_index, column=5, value=total_col_1)
        cell_val_1.font = total_red_font_style if total_col_1 < 0 else total_font_style
        cell_val_1.number_format = '#,##0.00'
        
        cell_val_2 = ws_nova.cell(row=total_row_index, column=11, value=total_col_2)
        cell_val_2.font = total_red_font_style if total_col_2 < 0 else total_font_style
        cell_val_2.number_format = '#,##0.00'
        
        cell_diff = ws_nova.cell(row=total_row_index, column=12, value=total_diff)
        cell_diff.font = total_red_font_style if total_diff < 0 else total_font_style
        cell_diff.number_format = '#,##0.00'
        
        for cell in ws_nova[total_row_index]:
            cell.fill = total_fill_style
    
    num_cols = ws_nova.max_column
    for i in range(1, num_cols + 1):
        col_letter = ws_nova.cell(row=1, column=i).column_letter
        max_length = 0
        for cell in ws_nova[col_letter]:
            is_summary_or_total = (cell.font == summary_font or cell.font == total_font_style or cell.font == total_red_font_style)
            if cell.row > 1 and not is_summary_or_total:
                cell.font = data_font
            try:
                if i in [5, 11, 12] and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    if cell.value < 0:
                        cell.font = red_data_font if cell.font != summary_font else red_font
                char_count = len(str(cell.value or ""))
                if char_count > max_length:
                    max_length = char_count
            except: pass
        ws_nova.column_dimensions[col_letter].width = max_length + 3

def _preparar_dataframe(df_raw, col_config):
    colunas_essenciais = [col_config['cnpj'], col_config['nome'], col_config['valor']]
    colunas_extras = [col for col in [col_config.get('nota'), col_config.get('data_emissao'), col_config.get('num_contrato'), col_config.get('liquidacao')] if col]
    colunas_para_extrair = list(dict.fromkeys(colunas_essenciais + colunas_extras))
    colunas_existentes = [col for col in colunas_para_extrair if col in df_raw.columns]
    if not all(c in df_raw.columns for c in colunas_essenciais):
        return pd.DataFrame()
    df = df_raw[colunas_existentes].copy()
    df.dropna(subset=[col_config['cnpj'], col_config['valor']], inplace=True)
    df.rename(columns={
        col_config['cnpj']: 'CNPJ_PADRAO',
        col_config['nome']: 'NOME_PADRAO',
        col_config['valor']: 'VALOR_PADRAO'
    }, inplace=True)
    df['CNPJ_PADRAO'] = df['CNPJ_PADRAO'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.zfill(14)
    df['valor_arredondado'] = pd.to_numeric(df['VALOR_PADRAO'], errors='coerce').round(2)
    df.dropna(subset=['valor_arredondado'], inplace=True)
    df['indice_original'] = df.index
    return df

def _capturar_transacoes_combinadas(df1, df2, indices_combinados_1, indices_combinados_2, config):
    """Captura as transações que foram combinadas (excluídas) durante o processo de matching"""
    # Pegar as transações que foram combinadas
    transacoes_combinadas_1 = df1[df1['indice_original'].isin(indices_combinados_1)].copy()
    transacoes_combinadas_2 = df2[df2['indice_original'].isin(indices_combinados_2)].copy()
    
    if transacoes_combinadas_1.empty and transacoes_combinadas_2.empty:
        return pd.DataFrame(columns=config['nomes_colunas_saida'])
    
    # Agrupar por chave de 8 dígitos para criar pares
    grouping_key_length = 8
    
    if not transacoes_combinadas_1.empty:
        transacoes_combinadas_1['chave_agrupamento'] = transacoes_combinadas_1['CNPJ_PADRAO'].str[:grouping_key_length]
        transacoes_combinadas_1['chave_emparelhamento'] = transacoes_combinadas_1.groupby('chave_agrupamento').cumcount()
    if not transacoes_combinadas_2.empty:
        transacoes_combinadas_2['chave_agrupamento'] = transacoes_combinadas_2['CNPJ_PADRAO'].str[:grouping_key_length]
        transacoes_combinadas_2['chave_emparelhamento'] = transacoes_combinadas_2.groupby('chave_agrupamento').cumcount()

    # Renomear colunas para merge
    cols_para_renomear_1 = {c: f"{c}_esq" for c in transacoes_combinadas_1.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    transacoes_combinadas_1.rename(columns=cols_para_renomear_1, inplace=True)
    cols_para_renomear_2 = {c: f"{c}_dir" for c in transacoes_combinadas_2.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    transacoes_combinadas_2.rename(columns=cols_para_renomear_2, inplace=True)

    # Fazer merge das transações combinadas
    merged_combinadas = pd.merge(
        transacoes_combinadas_1, transacoes_combinadas_2,
        on=['chave_agrupamento', 'chave_emparelhamento'],
        how='outer'
    )
    
    if merged_combinadas.empty:
        return pd.DataFrame(columns=config['nomes_colunas_saida'])
    
    return _criar_resultado_final(merged_combinadas, config, config['nomes_colunas_saida'])

def _processar_comparacao(df1, df2, config, indices_combinados_1, indices_combinados_2):
    nomes_saida = config['nomes_colunas_saida']
    sobras_1 = df1[~df1['indice_original'].isin(indices_combinados_1)].copy()
    sobras_2 = df2[~df2['indice_original'].isin(indices_combinados_2)].copy()
    
    # Capturar transações que foram combinadas (excluídas)
    transacoes_excluidas = _capturar_transacoes_combinadas(df1, df2, indices_combinados_1, indices_combinados_2, config)
    
    if sobras_1.empty and sobras_2.empty:
        return pd.DataFrame(columns=nomes_saida), transacoes_excluidas

    grouping_key_length = config.get('chave_agrupamento_final', 8)
    
    if not sobras_1.empty:
        sobras_1['chave_agrupamento'] = sobras_1['CNPJ_PADRAO'].str[:grouping_key_length]
        sobras_1['chave_emparelhamento'] = sobras_1.groupby('chave_agrupamento').cumcount()
    if not sobras_2.empty:
        sobras_2['chave_agrupamento'] = sobras_2['CNPJ_PADRAO'].str[:grouping_key_length]
        sobras_2['chave_emparelhamento'] = sobras_2.groupby('chave_agrupamento').cumcount()

    cols_para_renomear_1 = {c: f"{c}_esq" for c in sobras_1.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    sobras_1.rename(columns=cols_para_renomear_1, inplace=True)
    cols_para_renomear_2 = {c: f"{c}_dir" for c in sobras_2.columns if c not in ['chave_agrupamento', 'chave_emparelhamento']}
    sobras_2.rename(columns=cols_para_renomear_2, inplace=True)

    merged_sobras = pd.merge(
        sobras_1, sobras_2,
        on=['chave_agrupamento', 'chave_emparelhamento'],
        how='outer'
    )
    
    blocos_excluidos = pd.DataFrame()  # Para armazenar blocos excluídos por diferença pequena
    
    if not merged_sobras.empty:
        # Lógica de filtragem por diferença de valores
        group_sums = merged_sobras.groupby('chave_agrupamento')[['VALOR_PADRAO_esq', 'VALOR_PADRAO_dir']].sum()
        group_sums['diferenca_bloco'] = group_sums['VALOR_PADRAO_esq'] - group_sums['VALOR_PADRAO_dir']
        
        # Separar blocos que serão excluídos (diferença ≤ 0.01)
        chaves_excluidas = group_sums[abs(group_sums['diferenca_bloco']) <= 0.01].index
        blocos_excluidos = merged_sobras[merged_sobras['chave_agrupamento'].isin(chaves_excluidas)].copy()
        
        # Manter apenas blocos com diferença > 0.01
        chaves_para_manter = group_sums[abs(group_sums['diferenca_bloco']) > 0.01].index
        merged_sobras = merged_sobras[merged_sobras['chave_agrupamento'].isin(chaves_para_manter)].copy()

        if merged_sobras.empty:
            merged_sobras = pd.DataFrame(columns=nomes_saida)

    # Processar dados principais
    resultado_principal = _criar_resultado_final(merged_sobras, config, nomes_saida)
    
    # Processar blocos excluídos por diferença pequena
    resultado_blocos_excluidos = _criar_resultado_final(blocos_excluidos, config, nomes_saida) if not blocos_excluidos.empty else pd.DataFrame(columns=nomes_saida)
    
    # Combinar ambos os tipos de exclusões
    todos_excluidos = pd.concat([transacoes_excluidas, resultado_blocos_excluidos], ignore_index=True)
    
    return resultado_principal, todos_excluidos

def _criar_resultado_final(merged_data, config, nomes_saida):
    """Função auxiliar para criar o DataFrame final a partir dos dados merged"""
    if merged_data.empty:
        return pd.DataFrame(columns=nomes_saida)
        
    merged_data['Nome_Ordenacao'] = merged_data['NOME_PADRAO_esq'].fillna(merged_data['NOME_PADRAO_dir'])
    merged_data = merged_data.sort_values(
        by=['Nome_Ordenacao', 'chave_agrupamento', 'chave_emparelhamento']
    ).reset_index(drop=True)

    final_data = {}
    def safe_get_col(df, col_name):
        return df[col_name] if col_name and col_name in df.columns else pd.Series(index=df.index)
    
    # A coluna de diferença vai vazia nos dados - será preenchida na linha de resumo
    merged_data['diferenca_bloco'] = np.nan

    col_nota_esq = f"{config['colunas_aba_1'].get('nota')}_esq" if config['colunas_aba_1'].get('nota') else None
    col_data_esq = f"{config['colunas_aba_1'].get('data_emissao')}_esq" if config['colunas_aba_1'].get('data_emissao') else None
    final_data[nomes_saida[0]] = safe_get_col(merged_data, 'CNPJ_PADRAO_esq')
    final_data[nomes_saida[1]] = safe_get_col(merged_data, col_nota_esq)
    final_data[nomes_saida[2]] = safe_get_col(merged_data, col_data_esq)
    final_data[nomes_saida[3]] = safe_get_col(merged_data, 'NOME_PADRAO_esq')
    final_data[nomes_saida[4]] = safe_get_col(merged_data, 'VALOR_PADRAO_esq')
    final_data[nomes_saida[5]] = ''
    col_contrato_dir = f"{config['colunas_aba_2'].get('num_contrato')}_dir" if config['colunas_aba_2'].get('num_contrato') else None
    col_liq_dir = f"{config['colunas_aba_2'].get('liquidacao')}_dir" if config['colunas_aba_2'].get('liquidacao') else None
    final_data[nomes_saida[6]] = safe_get_col(merged_data, 'CNPJ_PADRAO_dir')
    final_data[nomes_saida[7]] = safe_get_col(merged_data, col_contrato_dir)
    final_data[nomes_saida[8]] = safe_get_col(merged_data, col_liq_dir)
    final_data[nomes_saida[9]] = safe_get_col(merged_data, 'NOME_PADRAO_dir')
    final_data[nomes_saida[10]] = safe_get_col(merged_data, 'VALOR_PADRAO_dir')
    final_data[nomes_saida[11]] = safe_get_col(merged_data, 'diferenca_bloco')
    
    return pd.DataFrame(final_data)

def _encontrar_melhores_matches(df1, df2, grouping_key_length):
    """
    Encontra os melhores pares de transações com base na menor diferença de valor,
    dentro de cada grupo de CNPJ de 8 dígitos.
    """
    df1['chave_agrupamento'] = df1['CNPJ_PADRAO'].str[:grouping_key_length]
    df2['chave_agrupamento'] = df2['CNPJ_PADRAO'].str[:grouping_key_length]
    
    todas_chaves = set(df1['chave_agrupamento'].dropna()).union(set(df2['chave_agrupamento'].dropna()))
    
    matched_indices_1 = set()
    matched_indices_2 = set()
    
    for chave in todas_chaves:
        grupo_esq = df1[df1['chave_agrupamento'] == chave].copy()
        grupo_dir = df2[df2['chave_agrupamento'] == chave].copy()
        
        # Itera sobre as transações de um lado e busca o melhor par do outro
        for _, row_e in grupo_esq.iterrows():
            if row_e['indice_original'] in matched_indices_1:
                continue
                
            melhor_diferenca = np.inf
            melhor_par_dir_idx = None
            
            for _, row_d in grupo_dir.iterrows():
                if row_d['indice_original'] in matched_indices_2:
                    continue
                
                diferenca = abs(row_e['valor_arredondado'] - row_d['valor_arredondado'])
                
                if diferenca < melhor_diferenca:
                    melhor_diferenca = diferenca
                    melhor_par_dir_idx = row_d['indice_original']
            
            if melhor_diferenca <= 0.01 and melhor_par_dir_idx is not None:
                matched_indices_1.add(row_e['indice_original'])
                matched_indices_2.add(melhor_par_dir_idx)

    return matched_indices_1, matched_indices_2

def executar_comparacao_lado_a_lado(workbook, config):
    if workbook is None: return workbook
    try:
        df1_raw = pd.read_excel(config['arquivo_excel'], sheet_name=config['indice_aba_1'], skiprows=config['pular_linhas_1'], header=0)
        cutoff_date_str = config.get('data_corte')
        date_col_name = config['colunas_aba_1'].get('data_emissao')
        if cutoff_date_str and date_col_name and date_col_name in df1_raw.columns:
            try:
                cutoff_date = pd.to_datetime(cutoff_date_str)
                df1_raw[date_col_name] = pd.to_datetime(df1_raw[date_col_name], errors='coerce')
                original_rows = len(df1_raw)
                df1_raw = df1_raw[df1_raw[date_col_name] < cutoff_date].copy()
                print(f"[{config['nome_processo']}] Filtro de data aplicado. {original_rows - len(df1_raw)} linhas removidas.")
            except Exception as e:
                print(f"Aviso: Não foi possível aplicar o filtro de data. Erro: {e}")
        df2_raw = pd.read_excel(config['arquivo_excel'], sheet_name=config['indice_aba_2'], skiprows=config['pular_linhas_2'], header=0)
        df1 = _preparar_dataframe(df1_raw, config['colunas_aba_1'])
        df2 = _preparar_dataframe(df2_raw, config['colunas_aba_2'])
        if df1.empty and df2.empty:
            _formatar_aba_final(workbook, config, pd.DataFrame(columns=config['nomes_colunas_saida']))
            return workbook
        merged = pd.merge(df1, df2, on='CNPJ_PADRAO', how='outer', suffixes=('_1', '_2'))
        condicao_match = abs(merged['valor_arredondado_1'] - merged['valor_arredondado_2']) <= 0.01
        indices_combinados_1 = set(merged.loc[condicao_match, 'indice_original_1'].dropna())
        indices_combinados_2 = set(merged.loc[condicao_match, 'indice_original_2'].dropna())
        resultado_final, _ = _processar_comparacao(df1, df2, config, indices_combinados_1, indices_combinados_2)
        _formatar_aba_final(workbook, config, resultado_final)
        print(f"[{config['nome_processo']}] Processo concluído.")
        return workbook
    except Exception as e:
        print(f"Ocorreu um erro fatal no {config['nome_processo']}: {e}")
        return workbook

def executar_comparacao_com_exclusao_parcial(workbook, config):
    if workbook is None: return workbook
    print(f"\n--- INICIANDO {config['nome_processo']} ---")
    try:
        df1_raw = pd.read_excel(config['arquivo_excel'], sheet_name=config['indice_aba_1'], skiprows=config['pular_linhas_1'], header=0)
        cutoff_date_str = config.get('data_corte')
        date_col_name = config['colunas_aba_1'].get('data_emissao')
        if cutoff_date_str and date_col_name and date_col_name in df1_raw.columns:
            try:
                cutoff_date = pd.to_datetime(cutoff_date_str)
                df1_raw[date_col_name] = pd.to_datetime(df1_raw[date_col_name], errors='coerce')
                original_rows = len(df1_raw)
                df1_raw = df1_raw[df1_raw[date_col_name] < cutoff_date].copy()
                print(f"[{config['nome_processo']}] Filtro de data aplicado. {original_rows - len(df1_raw)} linhas removidas.")
            except Exception as e:
                print(f"Aviso: Não foi possível aplicar o filtro de data. Erro: {e}")
        df2_raw = pd.read_excel(config['arquivo_excel'], sheet_name=config['indice_aba_2'], skiprows=config['pular_linhas_2'], header=0)
        
        df1 = _preparar_dataframe(df1_raw, config['colunas_aba_1'])
        df2 = _preparar_dataframe(df2_raw, config['colunas_aba_2'])
        if df1.empty and df2.empty:
            _formatar_aba_final(workbook, config, pd.DataFrame(columns=config['nomes_colunas_saida']))
            return workbook
            
        # Nova lógica de casamento de transações
        indices_combinados_1, indices_combinados_2 = _encontrar_melhores_matches(
            df1, df2, config.get('chave_agrupamento_final', 8)
        )
        
        resultado_final, resultado_excluidos = _processar_comparacao(df1, df2, config, indices_combinados_1, indices_combinados_2)
        
        # Criar a aba principal
        _formatar_aba_final(workbook, config, resultado_final)
        
        # Criar a aba dos dados excluídos se houver dados
        if not resultado_excluidos.empty:
            config_excluidos = config.copy()
            config_excluidos['nome_aba_saida'] = config['nome_aba_saida'] + ' - ='
            _formatar_aba_final(workbook, config_excluidos, resultado_excluidos)
            print(f"[{config['nome_processo']}] Aba de excluídos criada: {config_excluidos['nome_aba_saida']}")
        
        print(f"[{config['nome_processo']}] Processo concluído.")
        return workbook
    except Exception as e:
        print(f"Ocorreu um erro fatal no {config['nome_processo']}: {e}")
        return workbook