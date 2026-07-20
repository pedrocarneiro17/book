import os
import io
import functools
from dotenv import load_dotenv
load_dotenv()

import pandas as pd
import openpyxl
from flask import (
    Flask, render_template, request, send_file,
    make_response, redirect, url_for, session
)

from modelo1.parte1_processador import executar_processo_parte1
from modelo1.parte2_processador import executar_comparacao_lado_a_lado, executar_comparacao_com_exclusao_parcial
from modelo1.resumo_processador import criar_aba_resumo

from thunders.parte2_processador import (
    consolidar_books,
    executar_comparacao_thunders,
    executar_exclusao_parcial_thunders,
    executar_confronto_book_x_book,
    COLUNAS_LIVRO_SAIDA,
    NOMES_COLUNAS_SAIDA_SAIDA,
)
from thunders.resumo_processador import criar_aba_resumo_thunders

from zeus.parte2_processador import (
    consolidar_book as consolidar_book_zeus,
    executar_comparacao_zeus,
    executar_exclusao_parcial_zeus,
    COLUNAS_LIVRO_SAIDA as COLUNAS_LIVRO_SAIDA_ZEUS,
    NOMES_COLUNAS_SAIDA_SAIDA as NOMES_COLUNAS_SAIDA_SAIDA_ZEUS,
)
from zeus.resumo_processador import criar_aba_resumo_zeus

import auth

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-in-prod')

# Inicializa BD e cria master na primeira execução
try:
    auth.init_db()
except Exception as e:
    print(f"[AUTH] Aviso: não foi possível conectar à BD: {e}")


# ---------------------------------------------------------------------------
# Decoradores
# ---------------------------------------------------------------------------

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def master_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not session.get('is_master'):
            return "Acesso negado.", 403
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Login / Logout
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = auth.verify_login(username, password)
        if user:
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['is_master'] = user['is_master']
            # Regista IP
            ip = request.headers.get('X-Forwarded-For', request.remote_addr)
            if ip:
                ip = ip.split(',')[0].strip()
            ua = request.headers.get('User-Agent', '')
            try:
                auth.log_ip(user['id'], ip, ua)
            except Exception:
                pass
            return redirect(url_for('index'))
        else:
            error = "Username ou password incorretos."
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Painel Master
# ---------------------------------------------------------------------------

@app.route('/master')
@master_required
def master_panel():
    users      = auth.get_all_users()
    ip_rows    = auth.get_ip_summary()
    msg      = session.pop('flash_msg', None)
    msg_type = session.pop('flash_type', 'success')

    # Ajusta horário para UTC-3 (Brasil) e agrupa IPs por username
    from collections import OrderedDict
    from datetime import timedelta
    BR = timedelta(hours=-3)

    ip_por_user = OrderedDict()
    for row in ip_rows:
        if row['last_access']:
            row['last_access'] = row['last_access'] + BR
        name = row['username']
        if name not in ip_por_user:
            ip_por_user[name] = {'is_active': row['is_active'], 'ips': []}
        ip_por_user[name]['ips'].append(row)

    return render_template('master_panel.html',
                           users=users, ip_por_user=ip_por_user,
                           current_user_id=session['user_id'],
                           msg=msg, msg_type=msg_type)


@app.route('/master/create_user', methods=['POST'])
@master_required
def master_create_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    ok, err = auth.create_user(username, password)
    if ok:
        session['flash_msg']  = f"Utilizador '{username}' criado com sucesso."
        session['flash_type'] = 'success'
    else:
        session['flash_msg']  = f"Erro: {err}"
        session['flash_type'] = 'danger'
    return redirect(url_for('master_panel'))


@app.route('/master/toggle/<int:user_id>', methods=['POST'])
@master_required
def master_toggle_user(user_id):
    auth.toggle_user_active(user_id)
    return redirect(url_for('master_panel'))


@app.route('/master/reset_password', methods=['POST'])
@master_required
def master_reset_password():
    user_id      = request.form.get('user_id', type=int)
    new_password = request.form.get('new_password', '')
    if user_id and new_password:
        auth.update_password(user_id, new_password)
        session['flash_msg']  = "Password atualizada com sucesso."
        session['flash_type'] = 'success'
    return redirect(url_for('master_panel'))


@app.route('/master/clear_ips/<int:user_id>', methods=['POST'])
@master_required
def master_clear_ips(user_id):
    auth.clear_user_ips(user_id)
    session['flash_msg']  = "Histórico de IPs limpo."
    session['flash_type'] = 'success'
    return redirect(url_for('master_panel'))


@app.route('/master/delete/<int:user_id>', methods=['POST'])
@master_required
def master_delete_user(user_id):
    auth.delete_user(user_id, session['user_id'])
    session['flash_msg']  = "Utilizador apagado."
    session['flash_type'] = 'success'
    return redirect(url_for('master_panel'))


# ---------------------------------------------------------------------------
# App principal
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
@login_required
def process():
    if 'excel_file' not in request.files:
        return "Erro: Nenhum ficheiro foi enviado.", 400

    uploaded_file       = request.files['excel_file']
    modelo_selecionado  = request.form.get('modelo_selecionado')
    data_corte_compras  = request.form.get('data_corte_compras')
    data_corte_vendas   = request.form.get('data_corte_vendas')
    tipo_confronto_thunders = request.form.get('tipo_confronto_thunders', 'livro_book')
    data_corte_compras_bxb  = request.form.get('data_corte_compras_bxb')
    data_corte_vendas_bxb   = request.form.get('data_corte_vendas_bxb')

    if uploaded_file.filename == '':
        return "Erro: Nenhum ficheiro foi selecionado.", 400
    if not modelo_selecionado:
        return "Erro: Por favor, selecione um modelo de folha de cálculo.", 400

    if uploaded_file:
        try:
            print(f"Ficheiro recebido. A processar com o '{modelo_selecionado}'...")

            workbook_resultado = None

            if modelo_selecionado == 'modelo1':
                p1_pular_linhas_antes_cabecalho = 0
                p1_nomes_colunas = {
                    "nome": "Parte - Contra Banco", "cnpj": "CNPJ", "valor": "Valor Ajustado",
                    "deal": "Deal", "num_contrato": "Nº Contrato", "tipo_operacao": "Tipo de Operação",
                    "liquidacao": "Liquidação "
                }

                nomes_colunas_saida_unico   = ['CNPJ Esquerdo', 'Nota', 'Data Emissão', 'Fornecedor', 'Valor Contábil', ' ', 'CNPJ Direito', 'Nº Contrato', 'Liquidação', 'Parte - Contra Banco', 'Valor Ajustado', 'Diferença do Bloco']
                nomes_colunas_saida_cliente = ['CNPJ Esquerdo', 'Nota', 'Data Emissão', 'Cliente', 'Valor Contábil', ' ', 'CNPJ Direito', 'Nº Contrato', 'Liquidação', 'Parte - Contra Banco', 'Valor Ajustado', 'Diferença do Bloco']

                config_2_vs_5 = {
                    "nome_processo": "Parte 2.1 - Confronto Padrão", "arquivo_excel": uploaded_file, "indice_aba_1": 1, "indice_aba_2": 4,
                    "pular_linhas_1": 5, "pular_linhas_2": 0,
                    "nome_aba_saida": "Livro x Book Entrada - CNPJ",
                    "colunas_aba_1": {"nome": "Fornecedor", "cnpj": "CNPJ/CPF/CEI/CAEPF", "valor": "Valor Contábil", "nota": "Nota", "data_emissao": "Data Emissão"},
                    "colunas_aba_2": {"nome": "Parte - Contra Banco", "cnpj": "CNPJ", "valor": "Valor Ajustado", "num_contrato": "Nº Contrato", "liquidacao": "Liquidação "},
                    "nomes_colunas_saida": nomes_colunas_saida_unico,
                    "chave_agrupamento_final": 12,
                    "data_corte": data_corte_compras
                }
                config_1_vs_6 = {
                    "nome_processo": "Parte 2.2 - Confronto Padrão", "arquivo_excel": uploaded_file, "indice_aba_1": 0, "indice_aba_2": 5,
                    "pular_linhas_1": 5, "pular_linhas_2": 0,
                    "nome_aba_saida": "Livro x Book Saída - CNPJ",
                    "colunas_aba_1": {"nome": "Cliente", "cnpj": "CNPJ/CPF/CEI/CAEPF", "valor": "Valor Contábil", "nota": "Nota", "data_emissao": "Data Emissão"},
                    "colunas_aba_2": {"nome": "Parte - Contra Banco", "cnpj": "CNPJ", "valor": "Valor Ajustado", "num_contrato": "Nº Contrato", "liquidacao": "Liquidação "},
                    "nomes_colunas_saida": nomes_colunas_saida_cliente,
                    "chave_agrupamento_final": 12,
                    "data_corte": data_corte_vendas
                }
                config_parcial_2_vs_5 = {**config_2_vs_5,
                    "nome_processo": "Parte 3.1 - Confronto com Exclusão",
                    "nome_aba_saida": "Livro x Book Entrada",
                    "chave_agrupamento_final": 8
                }
                config_parcial_1_vs_6 = {**config_1_vs_6,
                    "nome_processo": "Parte 3.2 - Confronto com Exclusão",
                    "nome_aba_saida": "Livro x Book Saída",
                    "chave_agrupamento_final": 8
                }

                workbook_resultado = executar_processo_parte1(uploaded_file, p1_pular_linhas_antes_cabecalho, p1_nomes_colunas)
                workbook_resultado = executar_comparacao_lado_a_lado(workbook_resultado, config_2_vs_5)
                workbook_resultado = executar_comparacao_lado_a_lado(workbook_resultado, config_1_vs_6)
                workbook_resultado = executar_comparacao_com_exclusao_parcial(workbook_resultado, config_parcial_2_vs_5)
                workbook_resultado = executar_comparacao_com_exclusao_parcial(workbook_resultado, config_parcial_1_vs_6)

                if workbook_resultado:
                    workbook_resultado = criar_aba_resumo(
                        workbook_resultado, p1_nomes_colunas,
                        data_corte_compras, data_corte_vendas
                    )

            elif modelo_selecionado == 'thunders':
                file_bytes = uploaded_file.read()

                if tipo_confronto_thunders == 'book_book':
                    workbook_resultado = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_links=False, data_only=True)
                    workbook_resultado = executar_confronto_book_x_book(
                        workbook_resultado, io.BytesIO(file_bytes),
                        data_corte_compras=data_corte_compras_bxb,
                        data_corte_vendas=data_corte_vendas_bxb,
                    )

                else:
                    workbook_resultado   = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_links=False, data_only=True)
                    df_livro_entrada_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, skiprows=5, header=0)
                    df_livro_saida_raw   = pd.read_excel(io.BytesIO(file_bytes), sheet_name=1, skiprows=5, header=0)
                    df_book_compra, df_book_venda = consolidar_books(io.BytesIO(file_bytes))

                    config_entrada_cnpj = {
                        "nome_processo": "Thunders 2.1 - Livro x Book Entrada CNPJ",
                        "nome_aba_saida": "Livro x Book Entrada - CNPJ",
                        "chave_agrupamento_final": 12,
                        "data_corte": data_corte_compras,
                    }
                    config_saida_cnpj = {
                        "nome_processo": "Thunders 2.2 - Livro x Book Saída CNPJ",
                        "nome_aba_saida": "Livro x Book Saída - CNPJ",
                        "chave_agrupamento_final": 12,
                        "data_corte": data_corte_vendas,
                        "colunas_livro": COLUNAS_LIVRO_SAIDA,
                        "nomes_colunas_saida": NOMES_COLUNAS_SAIDA_SAIDA,
                    }
                    config_entrada_parcial = {
                        **config_entrada_cnpj,
                        "nome_processo": "Thunders 3.1 - Livro x Book Entrada Exclusão",
                        "nome_aba_saida": "Livro x Book Entrada",
                        "chave_agrupamento_final": 8,
                    }
                    config_saida_parcial = {
                        **config_saida_cnpj,
                        "nome_processo": "Thunders 3.2 - Livro x Book Saída Exclusão",
                        "nome_aba_saida": "Livro x Book Saída",
                        "chave_agrupamento_final": 8,
                    }

                    workbook_resultado = executar_comparacao_thunders(
                        workbook_resultado, config_entrada_cnpj, df_livro_entrada_raw, df_book_compra)
                    workbook_resultado = executar_comparacao_thunders(
                        workbook_resultado, config_saida_cnpj, df_livro_saida_raw, df_book_venda)
                    workbook_resultado = executar_exclusao_parcial_thunders(
                        workbook_resultado, config_entrada_parcial, df_livro_entrada_raw, df_book_compra)
                    workbook_resultado = executar_exclusao_parcial_thunders(
                        workbook_resultado, config_saida_parcial, df_livro_saida_raw, df_book_venda)

                    if workbook_resultado:
                        workbook_resultado = criar_aba_resumo_thunders(
                            workbook_resultado, data_corte_compras, data_corte_vendas)

            elif modelo_selecionado == 'zeus':
                file_bytes = uploaded_file.read()

                workbook_resultado   = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_links=False, data_only=True)
                df_livro_entrada_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name='ENTRADAS', skiprows=5, header=0)
                df_livro_saida_raw   = pd.read_excel(io.BytesIO(file_bytes), sheet_name='SAÍDAS',   skiprows=5, header=0)
                df_book_compra, df_book_venda = consolidar_book_zeus(io.BytesIO(file_bytes))

                config_entrada_cnpj = {
                    "nome_processo":           "Zeus 2.1 - Livro x Book Entrada CNPJ",
                    "nome_aba_saida":           "Livro x Book Entrada - CNPJ",
                    "chave_agrupamento_final":  12,
                    "data_corte":               data_corte_compras,
                }
                config_saida_cnpj = {
                    "nome_processo":           "Zeus 2.2 - Livro x Book Saída CNPJ",
                    "nome_aba_saida":           "Livro x Book Saída - CNPJ",
                    "chave_agrupamento_final":  12,
                    "data_corte":               data_corte_vendas,
                    "colunas_livro":            COLUNAS_LIVRO_SAIDA_ZEUS,
                    "nomes_colunas_saida":      NOMES_COLUNAS_SAIDA_SAIDA_ZEUS,
                }
                config_entrada_parcial = {
                    **config_entrada_cnpj,
                    "nome_processo":           "Zeus 3.1 - Livro x Book Entrada Exclusão",
                    "nome_aba_saida":           "Livro x Book Entrada",
                    "chave_agrupamento_final":  8,
                }
                config_saida_parcial = {
                    **config_saida_cnpj,
                    "nome_processo":           "Zeus 3.2 - Livro x Book Saída Exclusão",
                    "nome_aba_saida":           "Livro x Book Saída",
                    "chave_agrupamento_final":  8,
                }

                workbook_resultado = executar_comparacao_zeus(
                    workbook_resultado, config_entrada_cnpj, df_livro_entrada_raw, df_book_compra)
                workbook_resultado = executar_comparacao_zeus(
                    workbook_resultado, config_saida_cnpj, df_livro_saida_raw, df_book_venda)
                workbook_resultado = executar_exclusao_parcial_zeus(
                    workbook_resultado, config_entrada_parcial, df_livro_entrada_raw, df_book_compra)
                workbook_resultado = executar_exclusao_parcial_zeus(
                    workbook_resultado, config_saida_parcial, df_livro_saida_raw, df_book_venda)

                if workbook_resultado:
                    workbook_resultado = criar_aba_resumo_zeus(
                        workbook_resultado, data_corte_compras, data_corte_vendas)

            else:
                return f"Erro: Modelo '{modelo_selecionado}' não reconhecido.", 400

            if workbook_resultado:
                print("Processamento concluído. A preparar ficheiro para download...")
                output = io.BytesIO()
                workbook_resultado.save(output)
                output.seek(0)

                response = make_response(send_file(
                    output, as_attachment=True,
                    download_name=uploaded_file.filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                ))
                response.set_cookie('fileDownload', 'true', max_age=60)
                return response
            else:
                return "Ocorreu um erro durante o processamento.", 500

        except Exception as e:
            print(f"Erro detalhado: {e}")
            return f"Ocorreu um erro inesperado durante o processamento: {e}", 500

    return "Erro desconhecido.", 500


if __name__ == '__main__':
    import sys
    sys.stdout.reconfigure(line_buffering=True)
    app.run(debug=True, port=8080, use_reloader=False)
